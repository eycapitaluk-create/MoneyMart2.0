#!/usr/bin/env python3
"""Investing/Yahoo 風 Price History xlsx: Date, Close, Volume を JSON で stdout へ。"""
import json
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def main():
    if len(sys.argv) < 2:
        print("Usage: extract-xlsx-close-volume.py /path/to/file.xlsx", file=sys.stderr)
        sys.exit(1)
    path = sys.argv[1]
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header = None
        header_row_idx = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if not row or row[0] is None:
                continue
            h0 = str(row[0]).strip().lower()
            if h0 == "date":
                header = [str(c).strip() if c is not None else "" for c in row]
                header_row_idx = i
                break
        if not header:
            print("No 'Date' header row found", file=sys.stderr)
            sys.exit(1)

        def col_exact(name: str) -> int:
            for j, h in enumerate(header):
                if h.lower() == name.lower():
                    return j
            return -1

        i_date = col_exact("Date")
        i_close = col_exact("Close")
        i_vol = col_exact("Volume")
        if i_date < 0 or i_close < 0:
            print("Need Date and Close columns", file=sys.stderr)
            sys.exit(1)

        out = []
        for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
            if not row or len(row) <= max(i_close, i_vol if i_vol >= 0 else i_close):
                continue
            d = row[i_date]
            if d is None:
                continue
            if hasattr(d, "strftime"):
                ds = d.strftime("%Y-%m-%d")
            else:
                ds = str(d).strip()[:10]
            if len(ds) < 10 or not ds[4] == "-":
                continue
            close_raw = row[i_close]
            try:
                c = float(close_raw) if close_raw is not None else None
            except (TypeError, ValueError):
                c = None
            v = None
            if i_vol >= 0:
                vr = row[i_vol]
                try:
                    v = float(vr) if vr is not None else None
                except (TypeError, ValueError):
                    v = None
            if c is None or c <= 0:
                continue
            out.append({"date": ds, "close": c, "volume": v})
    finally:
        wb.close()
    json.dump(out, sys.stdout)


if __name__ == "__main__":
    main()
